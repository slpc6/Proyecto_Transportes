from fastapi import APIRouter, HTTPException
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from model.interno import DatosInternoEnvio
from util.path import Path
from util.titulo import generar_titulo

router = APIRouter()

def crear_excel_con_formato(df_datos, titulo, output_file):
    """Función optimizada para crear Excel con formato"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Título
    ws['A1'] = titulo
    ws.merge_cells('A1:F1')  # Sin columna destino
    titulo_cell = ws['A1']
    titulo_cell.font = Font(bold=True, size=14)
    titulo_cell.alignment = Alignment(horizontal='center')
    
    # Encabezados
    headers = ['Fecha', 'Placa', '# de Viajes', 'Tipo de Vehículo', 'Valor Unitario', 'Valor Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    for row_idx, row in enumerate(dataframe_to_rows(df_datos, index=False, header=False), 3):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Formato de moneda para columnas de valores
            if col_idx in [5, 6]:  # Valor Unitario y Valor Total
                cell.number_format = '"$"#,##0.00'
    
    # Total
    total_fila = len(df_datos) + 3
    ws.cell(row=total_fila, column=4, value="TOTAL").font = Font(bold=True)  # Tipo de Vehículo
    total_valor = df_datos['Valor Total'].sum()
    total_cell = ws.cell(row=total_fila, column=6, value=total_valor)  # Valor Total
    total_cell.font = Font(bold=True)
    total_cell.number_format = '"$"#,##0.00'
    
    # Anchos de columnas optimizados
    column_widths = {'A': 12, 'B': 10, 'C': 12, 'D': 20, 'E': 15, 'F': 15}
    for column_letter, width in column_widths.items():
        ws.column_dimensions[column_letter].width = width
    
    wb.save(output_file)
    return total_valor

@router.post("/interno")
def interno(request: DatosInternoEnvio):
    try:
        input_file = os.path.join(Path.INPUT, "INTERNOS.xlsx")
        output_file = os.path.join(Path.OUTPUT, "INTERNOS.xlsx")
        
        if not os.path.exists(input_file):
            raise HTTPException(status_code=404, detail="Archivo INTERNOS.xlsx no encontrado en input/")
        
        os.makedirs(Path.OUTPUT, exist_ok=True)
        
        # Leer datos existentes
        df_existente = pd.read_excel(input_file, header=1)
        
        # Preparar nuevos datos (sin destino)
        nuevos_datos = [{
            'Fecha': registro.fecha,
            'Placa': registro.placa,
            '# de Viajes': registro.numViajes,
            'Tipo de Vehículo': registro.tipoVehiculo,
            'Valor Unitario': registro.valorUnitario,
            'Valor Total': registro.valorTotal
        } for registro in request.registros]
        
        df_final = pd.DataFrame(nuevos_datos)
        titulo = generar_titulo('interno')
        
        # Crear Excel optimizado
        total_valor = crear_excel_con_formato(df_final, titulo, output_file)
        
        return {
            "success": True,
            "message": f"Archivo guardado exitosamente en {output_file}",
            "registros_procesados": len(request.registros),
            "total": request.total,
            "fecha_procesamiento": datetime.now().isoformat(),
            "titulo_generado": titulo,
            "total_calculado": float(total_valor)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al procesar los datos: {str(e)}") 