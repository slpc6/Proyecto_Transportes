from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import DataFrame
import os

from util.path import Path


def generar_archivo(df_final: DataFrame, 
                    titulo: str, 
                    headers: list[str], 
                    nombre: str = 'ARCHIVO')-> int:
    """Metodo para generar cada Relacion de forma generica
    :args:
    - df_final: Dataframe con la informacion que sera guardada como archivo excel.
    - titulo: Primera linea del datraframe que registra el titulo que tendra el documento.
    - Headers: Losta con los nombre de las columnas del DataFrame.
    - Nombre: Nombre que tendra del archivo excel.

    :returns:
    - EL valor calculado de sumar la ultima columna del DataFrame.

    """
    output_file = os.path.join(Path.OUTPUT, f"{nombre}.xlsx")
    os.makedirs(Path.OUTPUT, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Relacion"
        
    ws['A1'] = titulo
    ws.merge_cells('A1:G1')
    titulo_cell = ws['A1']
    titulo_cell.font = Font(bold=True, size=14)
    titulo_cell.alignment = Alignment(horizontal='center')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(dataframe_to_rows(df_final, index=False, header=False), 3):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 6:  # Valor Unitario
                cell.number_format = '"$"#,##0.00'
            elif col_idx == 7:  # Valor Total
                cell.number_format = '"$"#,##0.00'
        
        
    total_fila = len(df_final) + 3
        
       
    ws.cell(row=total_fila, column=5, value="TOTAL").font = Font(bold=True)
        
    total_valor = df_final['Valor Total'].sum()
    total_cell = ws.cell(row=total_fila, column=7, value=total_valor)
    total_cell.font = Font(bold=True)
    total_cell.number_format = '"$"#,##0.00'
        
    column_widths = {
            'A': 12,  # Fecha
            'B': 10,  # Placa
            'C': 20,  # Destino
            'D': 12,  # # de Viajes
            'E': 20,  # Tipo de Vehículo
            'F': 15,  # Valor Unitario
            'G': 15   # Valor Total
        }
        
    for column_letter, width in column_widths.items():
            ws.column_dimensions[column_letter].width = width
        
    wb.save(output_file)
     
    return total_valor
